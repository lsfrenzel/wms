from flask import render_template, jsonify, Response
from flask_login import login_required
from routes import reports_bp
from models import User, Product, Movement
from datetime import datetime, timedelta
import random
import csv
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

@reports_bp.route('/')
@login_required
def index():
    return render_template('reports.html')

@reports_bp.route('/api/user_stats')
@login_required
def user_stats():
    total_users = User.query.count()
    active_users = User.query.filter(User.active.is_(True)).count()
    inactive_users = total_users - active_users
    admin_users = User.query.filter_by(role='admin').count()
    regular_users = User.query.filter_by(role='user').count()
    
    return jsonify({
        'total': total_users,
        'active': active_users,
        'inactive': inactive_users,
        'admins': admin_users,
        'regular': regular_users
    })

@reports_bp.route('/api/stock_movements')
@login_required
def stock_movements():
    days = 7
    labels = []
    entries = []
    exits = []
    
    for i in range(days):
        date = datetime.now() - timedelta(days=days-i-1)
        labels.append(date.strftime('%d/%m'))
        entries.append(random.randint(10, 50))
        exits.append(random.randint(5, 45))
    
    return jsonify({
        'labels': labels,
        'entries': entries,
        'exits': exits
    })

@reports_bp.route('/api/stock_by_category')
@login_required
def stock_by_category():
    categories = ['Eletrônicos', 'Roupas', 'Alimentos', 'Móveis', 'Ferramentas', 'Outros']
    quantities = [random.randint(20, 100) for _ in range(len(categories))]
    
    return jsonify({
        'labels': categories,
        'data': quantities
    })

@reports_bp.route('/api/recent_activities')
@login_required
def recent_activities():
    activities = [
        {'type': 'Entrada', 'item': 'Notebook Dell XPS 15', 'quantity': 10, 'date': '2024-11-01 10:30'},
        {'type': 'Saída', 'item': 'Mouse Logitech MX Master', 'quantity': 5, 'date': '2024-11-01 09:15'},
        {'type': 'Entrada', 'item': 'Teclado Mecânico RGB', 'quantity': 15, 'date': '2024-10-31 16:45'},
        {'type': 'Saída', 'item': 'Monitor LG 27"', 'quantity': 3, 'date': '2024-10-31 14:20'},
        {'type': 'Entrada', 'item': 'Webcam Logitech C920', 'quantity': 8, 'date': '2024-10-31 11:00'},
        {'type': 'Ajuste', 'item': 'Headset HyperX Cloud', 'quantity': -2, 'date': '2024-10-30 15:30'},
    ]
    
    return jsonify(activities)

@reports_bp.route('/export/csv')
@login_required
def export_csv():
    output = StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Relatório de Atividades - WMS Sistema'])
    writer.writerow(['Gerado em:', datetime.now().strftime('%d/%m/%Y %H:%M')])
    writer.writerow([])
    
    writer.writerow(['Estatísticas de Usuários'])
    writer.writerow(['Tipo', 'Quantidade'])
    writer.writerow(['Total de Usuários', User.query.count()])
    writer.writerow(['Usuários Ativos', User.query.filter_by(active=True).count()])
    writer.writerow(['Administradores', User.query.filter_by(role='admin').count()])
    writer.writerow([])
    
    writer.writerow(['Atividades Recentes'])
    writer.writerow(['Tipo', 'Item', 'Quantidade', 'Data/Hora'])
    
    activities = [
        ['Entrada', 'Notebook Dell XPS 15', 10, '2024-11-01 10:30'],
        ['Saída', 'Mouse Logitech MX Master', 5, '2024-11-01 09:15'],
        ['Entrada', 'Teclado Mecânico RGB', 15, '2024-10-31 16:45'],
        ['Saída', 'Monitor LG 27"', 3, '2024-10-31 14:20'],
        ['Entrada', 'Webcam Logitech C920', 8, '2024-10-31 11:00'],
        ['Ajuste', 'Headset HyperX Cloud', -2, '2024-10-30 15:30'],
    ]
    
    for activity in activities:
        writer.writerow(activity)
    
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_wms_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    return response

@reports_bp.route('/export/xlsx')
@login_required
def export_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório WMS"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    ws['A1'] = 'Relatório de Atividades - WMS Sistema'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws.merge_cells('A2:D2')
    ws['A2'].font = Font(italic=True)
    
    ws.append([])
    
    ws.append(['Estatísticas de Usuários'])
    ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
    
    headers = ['Tipo', 'Quantidade']
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    ws.append(['Total de Usuários', User.query.count()])
    ws.append(['Usuários Ativos', User.query.filter_by(active=True).count()])
    ws.append(['Administradores', User.query.filter_by(role='admin').count()])
    ws.append(['Usuários Regulares', User.query.filter_by(role='user').count()])
    
    ws.append([])
    
    ws.append(['Produtos em Estoque'])
    ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
    
    product_headers = ['Código', 'Nome', 'Categoria', 'Quantidade', 'Localização']
    ws.append(product_headers)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    products = Product.query.all()
    for product in products:
        ws.append([
            product.code,
            product.name,
            product.category,
            product.quantity,
            product.location or 'N/A'
        ])
    
    ws.append([])
    
    ws.append(['Atividades Recentes'])
    ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
    
    activity_headers = ['Tipo', 'Item', 'Quantidade', 'Data/Hora']
    ws.append(activity_headers)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    activities = [
        ['Entrada', 'Notebook Dell XPS 15', 10, '2024-11-01 10:30'],
        ['Saída', 'Mouse Logitech MX Master', 5, '2024-11-01 09:15'],
        ['Entrada', 'Teclado Mecânico RGB', 15, '2024-10-31 16:45'],
        ['Saída', 'Monitor LG 27"', 3, '2024-10-31 14:20'],
        ['Entrada', 'Webcam Logitech C920', 8, '2024-10-31 11:00'],
        ['Ajuste', 'Headset HyperX Cloud', -2, '2024-10-30 15:30'],
    ]
    
    for activity in activities:
        ws.append(activity)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = f'attachment; filename=relatorio_wms_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

@reports_bp.route('/export/pdf')
@login_required
def export_pdf():
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    elements.append(Paragraph('Relatório de Gráficos - WMS Sistema', title_style))
    elements.append(Paragraph(f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    elements.append(Paragraph('1. Estatísticas de Usuários', heading_style))
    
    total_users = User.query.count()
    active_users = User.query.filter(User.active.is_(True)).count()
    inactive_users = total_users - active_users
    admin_users = User.query.filter_by(role='admin').count()
    regular_users = User.query.filter_by(role='user').count()
    
    user_data = [
        ['Métrica', 'Quantidade'],
        ['Total de Usuários', str(total_users)],
        ['Usuários Ativos', str(active_users)],
        ['Usuários Inativos', str(inactive_users)],
        ['Administradores', str(admin_users)],
        ['Usuários Regulares', str(regular_users)],
    ]
    
    user_table = Table(user_data, colWidths=[3.5*inch, 2*inch])
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(user_table)
    elements.append(Spacer(1, 0.3*inch))
    
    elements.append(Paragraph('2. Estoque por Categoria', heading_style))
    
    categories = ['Eletrônicos', 'Roupas', 'Alimentos', 'Móveis', 'Ferramentas', 'Outros']
    quantities = [random.randint(20, 100) for _ in range(len(categories))]
    
    category_data = [['Categoria', 'Quantidade em Estoque']]
    for cat, qty in zip(categories, quantities):
        category_data.append([cat, str(qty)])
    
    category_table = Table(category_data, colWidths=[3.5*inch, 2*inch])
    category_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(category_table)
    elements.append(Spacer(1, 0.3*inch))
    
    elements.append(Paragraph('3. Movimentações de Estoque (Últimos 7 dias)', heading_style))
    
    days = 7
    movement_data = [['Data', 'Entradas', 'Saídas']]
    for i in range(days):
        date = datetime.now() - timedelta(days=days-i-1)
        entries = random.randint(10, 50)
        exits = random.randint(5, 45)
        movement_data.append([
            date.strftime('%d/%m'),
            str(entries),
            str(exits)
        ])
    
    movement_table = Table(movement_data, colWidths=[2*inch, 1.75*inch, 1.75*inch])
    movement_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(movement_table)
    elements.append(Spacer(1, 0.3*inch))
    
    elements.append(Paragraph('4. Produtos em Estoque', heading_style))
    
    products = Product.query.limit(10).all()
    product_data = [['Código', 'Nome', 'Quantidade', 'Localização']]
    
    for product in products:
        product_data.append([
            product.code,
            product.name[:30] + '...' if len(product.name) > 30 else product.name,
            str(product.quantity),
            product.location or 'N/A'
        ])
    
    if not products:
        product_data.append(['Nenhum produto cadastrado', '', '', ''])
    
    product_table = Table(product_data, colWidths=[1.2*inch, 2.8*inch, 1*inch, 1*inch])
    product_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(product_table)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = Response(buffer.getvalue(), mimetype='application/pdf')
    response.headers['Content-Disposition'] = f'attachment; filename=graficos_wms_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response
